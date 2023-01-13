import json
import sys
from collections import Counter
from typing import Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

import tools as t


def update_contents(
    ws: Worksheet, temp: t.type_contents, now: t.datetime
) -> t.type_contents:
    """Update contents.json and contents tab of excel loader file. """
    
    for date, *row in ws:
        if not date.value:
            row = t.get_cells_values(row)
            
            # ignore incomplete record
            if any(val is None for val in row):
                date.value = 'ValueError'
                continue
            
            item, elem, pcs = row
                
            if type(pcs) is not int:
                date.value = 'TypeError'
                continue
            
            if pcs > 0:
                temp.setdefault(item, {}).update({elem: pcs})
                date.value = now
            elif not pcs:  # zero qty works as deletion
                try:
                    del temp[item][elem]
                    if not temp[item]:
                        del temp[item]
                    date.value = now
                except KeyError:
                    date.value = 'KeyError'
            else:  # negative qty 
                date.value = 'ValueError'
                
    subitems = {}  # elems that are a part of _another_ items 
    
    for item, elems in temp.items():
        for elem in elems:
            if elem != item:  # exclude monoitems and looped references 
                subitems.setdefault(elem, []).append(item)
                
    path = t.dumps_dir / 'subitems.json'
    
    with path.open('w', encoding='windows-1251') as target:
            json.dump(subitems, target, ensure_ascii=False, indent=4)
                
    return temp


def update_preforms(
    ws: Worksheet, temp: t.type_preforms, now: t.datetime
) -> t.type_preforms:
    """Update preforms.json and contents tab of excel loader file. """
    
    HEADERS = 'a', 'b', 'sheet'
    
    for date, *row in ws:
        if not date.value:
            row = t.get_cells_values(row)
            
            if any(val is None for val in row):
                date.value = 'ValueError'
                continue
            
            elem, *shape, sheet = row
            
            for x in shape:
                if type(x) not in (int, float):
                    date.value = 'TypeError'
                    break
                if x <= 0:
                    date.value = 'ValueError'
                    break
            else:
                temp[elem] = dict(zip(HEADERS, shape + [sheet]))
                date.value = now
    
    return temp


def update_feedstock(
    ws: Worksheet, temp: t.type_contents, now: t.datetime
) -> t.type_contents:
    """Update feedstock.json and contents tab of excel loader file. """
    
    HEADERS = 'd', 'W', 'L', 'rho'
    
    for date, *row in ws:
        if not date.value:
            row = t.get_cells_values(row)
            
            if any(val is None for val in row):
                date.value = 'ValueError'
                continue
            
            sheet, *params = row
                
            for p in params:
                if type(p) not in (int, float):
                    date.value = 'TypeError'
                    break
                if p <= 0:
                    date.value = 'ValueError'
                    break
            else:
                temp[sheet] = dict(zip(HEADERS, params))
                date.value = now
    
    return temp


def update_dump(tabs: list[str]) -> None:
    """Update json dump and excel loader tab with respective processor func. """
    
    funcs = update_contents, update_preforms, update_feedstock
    processors = dict(zip(t.holy_tabs, funcs))
    
    for tab in tabs:
        if tab not in processors:
            print(f'Wrong tab: {tab}')
            continue
        
        path = t.dumps_dir / f'{tab}.json'
        temp = t.read_dump(path) if path.exists() else {}
                
        wb = openpyxl.load_workbook('loader.xlsx')
        temp = processors[tab](wb[tab], temp, t.datetime.now())
        wb.save('loader.xlsx')
        
        temp = t.sort_nested_dict(temp)
        
        with path.open('w', encoding='windows-1251') as target:
            json.dump(temp, target, ensure_ascii=False, indent=4)
        

def calculate_sheets_number(grouped: Optional[list[str]]) -> None:
    """Calculate metal sheets number required to fulfill given items demand. """
    
    HEADER = ['date', 'sheet', 'mode', 'elem', 'items', 
              'demand, pcs', 'recomm, pcs', 'sheets_n']
    MSG = 'No demand to process or incorrect values in demand'
    
    temp, memo = {}, {}
    now = t.now_as_string()
    if grouped is None: 
        grouped = []
    
    wb = openpyxl.load_workbook('loader.xlsx')
    
    for date, *row in wb['demand']:
        if not date.value:
            item, qty = t.get_cells_values(row)
            
            if item not in t.contents:
                date.value = 'KeyError'
                continue
            if type(qty) is not int:
                date.value = 'TypeError'
                continue
            if qty <= 0:
                date.value = 'ValueError'
                continue
            
            date.value = now
            
            for elem, pcs in t.contents[item].items():
                a, b, sheet = t.preforms[elem].values()
                pps = (memo.setdefault(elem, {})
                       .setdefault('pps', t.get_pps(a, b, sheet, t.feedstock)))
                d = {elem: (dem := qty * pcs) / pps}
                mode = 'grouped' if t.has_marker(item, grouped) else 'single'
                temp.setdefault(sheet, {}).setdefault(mode, Counter()).update(d)
                memo[elem]['demand'] = memo[elem].get('demand', 0) + dem
                memo[elem].setdefault('items', []).append(item)
                
    temp = t.sort_nested_dict(temp)
    report = []
    
    for sheet, sheetdict in temp.items():
        for mode, modedict in sheetdict.items():
            
            # transform collected data mode-wise
            sheetdict[mode] = t.transformer[mode](modedict)
            
            # arrange transformed data to a report record
            for elem, sheets_n in temp[sheet][mode].items():
                items = ' '.join(memo[elem]['items'])
                demand = memo[elem]['demand']
                recomm = int(sheets_n * memo[elem]['pps'])
                report.append([now, sheet, mode, elem, items, 
                               demand, recomm, round(sheets_n, 1)])
    
    for row in report:
        wb['report'].append(row)
    
    wb.save('loader.xlsx')
    t.print_table(report, HEADER) if report else print(MSG)
    

def call_item(tester: list[str]) -> None:
    """Search for items containing any substring from tester list and 
    represent items' parameters. 
    """
    
    PARAMS_HEADER = ['N', 'item', 'elem', 'pcs',
                     'a, mm', 'b, mm', 'sheet', 'pps']    
    SUBITEMS_HEADER = ['N.n', 'item', 'elem', 'pcs']
    
    items = [item for item in t.contents if t.has_marker(item, tester)]
    
    if items:
        params_report, subitems_report = [], []
        memo, N = {}, 1
        
        for item in items:
            for elem, pcs in t.contents[item].items():
                params = t.preforms[elem].values()
                pps = memo.setdefault(elem, t.get_pps(*params, t.feedstock))
                params_report.append([N, item, elem, pcs, *params, pps])
            
            if item in t.subitems:  # elems that are a part of _another_ items
                n = 1  # create or reset subcounter
                for item_ in t.subitems[item]:
                    for elem_, pcs_ in t.contents[item_].items():
                        pointer = '>' if elem_ == item else ' '
                        row = [f'{pointer}{N}.{n}', item_, elem_, pcs_]
                        subitems_report.append(row)
                    n += 1 
            N += 1
        
        t.print_table(params_report, PARAMS_HEADER)
        
        if subitems_report:
            print('\n(N corresponds to N in table above)')
            t.print_table(subitems_report, SUBITEMS_HEADER)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = PARAMS_HEADER, SUBITEMS_HEADER
        reports = params_report, subitems_report
        
        for header, report in zip(headers, reports):
            if report:
                ws.append(header)
                for row in report:
                    ws.append(row)
                ws.append([])
        
        wb.save(t.calls_dir / f'{t.now_as_string()} call={tester}.xlsx')
    
    else: 
        print('Your call did not match any items')


if __name__ == '__main__':
    action, targets, grouped = t.get_command(sys.argv[1:])
    
    translator = {
        'go': calculate_sheets_number, 
        'call': call_item, 
        'update': update_dump 
    }
    
    if not targets:
        calculate_sheets_number(grouped)
    else: 
        translator[action](targets)
        